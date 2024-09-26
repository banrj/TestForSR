from uuid import UUID

from httpx import AsyncClient, HTTPError
from asyncio import get_event_loop
from openpyxl import load_workbook

from fastapi import UploadFile, HTTPException
from fastapi.responses import JSONResponse
from pydantic import ValidationError

from app.dto.book import (SearchBook, CreateBook, ShowBook,
                          DeleteBook, UpdateBook, InvalidBook,
                          )
from app.db.cruds import book as book_crud
from app.db.cruds import price_history as history_crud
from app.config import COLUMNS, LOAD_BOOKS_URL


async def create_new_book(body: CreateBook):
    new_book = await book_crud.create_book(
        title=body.title, author=body.author, genre=body.genre,
        description=body.description, cover_image=body.cover_image,
        publication_year=body.publication_year, price=body.price
    )
    await history_crud.create_history(book_id=new_book.id, price=new_book.price)
    return ShowBook(
        id=new_book.id, title=new_book.title, genre=new_book.genre,
        author=new_book.author, description=new_book.description,
        publication_year=new_book.publication_year, price=new_book.price,
        cover_image=new_book.cover_image
    )


async def delete_book(body: DeleteBook):
    return await book_crud.delete_book(book_id=body.id)


async def get_current_book(book_id: UUID):
    return await book_crud.get_book(book_id=book_id)


async def update_current_book(book_id: UUID, body: UpdateBook) -> ShowBook:
    if body.genre == ['string']:
        body.genre = None
    data = await book_crud.update_book(book_id=book_id, update_book=body)
    if data['flag']:
        await history_crud.create_history(book_id=book_id, price=body.price)
    return data['book']


async def get_list_books(
        lim: int, offset: int, title: str = None,
        author: str = None, genres: str = None, price: int = None,
        description: str = None, genres_neq: str = None
                        ):
    return await book_crud.get_books(
        lim=lim, offset=offset, title=title,
        author=author, genres=genres, genres_neq=genres_neq,
        description=description, price=price
    )


async def _make_api_request():
    async with AsyncClient() as client:
        try:
            response = await client.get(LOAD_BOOKS_URL)
            response.raise_for_status()
            data = response.json()
            return data
        except HTTPError as ex:
            return JSONResponse({'message': f'Возникла ошибка во время стороннего запроса!: {ex}'}, status_code=500)


async def filter_books():
    books_data = await _make_api_request()
    loaded_book = []
    unloaded_book = []
    for book in books_data:
        try:
            new = CreateBook(**book)
            loaded_book.append(new.json())
        except ValidationError:
            unloaded_book.append(book)
    information = await book_crud.load_data(loaded_book)

    return {'loaded_books': information['valid_books'],
            'invalid_books': unloaded_book, 'duplicates': information['duplicate_books']}


async def finalize_books(func, *args, **kwargs):
    books_data = await func(*args, **kwargs)

    price_data = [(book.id, book.price) for book in books_data['loaded_books']]
    await history_crud.create_many_history(price_data)
    return books_data


async def run_in_threadpool(func, *args, **kwargs):
    loop = get_event_loop()
    return await loop.run_in_executor(None, func, *args, **kwargs)


async def _validate_columns(actual_columns):
    if set(COLUMNS.keys()) != set(actual_columns):
        return JSONResponse(
            content={'Ваши поля не соответсвуют стандартной форме.\n'
                     'title|publication_year|genre|price|author|description|cover_image\n'
                     'Внимательно следите чтобы в названиях не было дополнительно символов и пробелов'},
            status_code=400)


async def check_file(file: UploadFile):
    if file.filename[-5:] != '.xlsx':
        raise HTTPException(detail={'message': 'Файл должен быть в формате .xlsx'}, status_code=400)

    try:
        workbook = await run_in_threadpool(load_workbook, file.file)
    except Exception as e:
        raise HTTPException(detail={'message': f'Ошибка чтения файла: {str(e)}'}, status_code=400)

    sheet = workbook.active

    columns = [cell.value for cell in sheet[1]]

    check_response = await _validate_columns(columns)
    if isinstance(check_response, JSONResponse):
        return check_response

    valid = []
    invalid = []

    # Проходим по каждой строке, начиная со второй (первая строка — заголовки)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        book = [
            row[0],  # title
            row[1],  # publication_year
            row[2],  # genre
            row[3],  # price
            row[4],  # author
            row[5],  # description
            row[6]   # cover_image
        ]

        if isinstance(row[2], str):
            book[2] = row[2].split(',')

        try:
            new = CreateBook(
                title=book[0],
                publication_year=book[1],
                genre=book[2],
                price=book[3],
                author=book[4],
                description=book[5],
                cover_image=book[6]
            )
            valid.append(new.json())
        except ValidationError as ex:
            invalid_book = InvalidBook(
                title=book[0],
                publication_year=book[1],
                genre=book[2],
                price=book[3],
                author=book[4],
                error=str(ex),
                description=book[5],
                cover_image=book[6]
            )
            invalid.append(invalid_book)

    information = await book_crud.load_data(valid)

    return {
        'loaded_books': information['valid_books'],
        'invalid_books': invalid,
        'duplicates': information['duplicate_books']
    }
